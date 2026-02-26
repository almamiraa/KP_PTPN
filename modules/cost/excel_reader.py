"""
excel_reader.py
===============
Module untuk membaca dan parse Excel file SDM
"""

import openpyxl
from pathlib import Path
from datetime import date
from typing import Dict, Tuple, Optional, List, Any
import logging
import re

logger = logging.getLogger(__name__)


class ExcelReader:
    
    # Mapping nama bulan di Excel (format yang digunakan di file)
    MONTH_MAP = {
        'jan': 1,
        'feb': 2,
        'mar': 3,
        'apr': 4,
        'mei': 5,
        'juni': 6,
        'juli': 7,
        'agst': 8,
        'sept': 9,
        'okt': 10,
        'nov': 11,
        'des': 12
    }
    
    # Row constants - posisi header di Excel
    PERIOD_HEADER_ROW = 7  # Row untuk "Desember 2024" | "s.d. Desember 2024"
    TYPE_HEADER_ROW = 8    # Row untuk "REAL 2024" | "RKAP 2024"
    
    def __init__(self, file_path: str):

        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet_names = []
        self._load_workbook()
    
    def _load_workbook(self) -> None:
        """Load Excel workbook"""
        try:
            # data_only=True untuk ambil nilai hasil formula, bukan formula-nya
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=True,
                read_only=False
            )
            
            self.sheet_names = self.workbook.sheetnames
            
            logger.info(f"Excel file loaded: {self.file_path.name}")
            logger.info(f"Found {len(self.sheet_names)} sheets")
            
            print(f"✓ Excel loaded: {self.file_path.name}")
            print(f"  Total sheets: {len(self.sheet_names)}")
            print(f"  Sheet names: {', '.join(self.sheet_names[:5])}" + 
                  (f"... (+{len(self.sheet_names)-5} more)" if len(self.sheet_names) > 5 else ""))
            
        except FileNotFoundError:
            error_msg = f"Excel file not found: {self.file_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        except Exception as e:
            error_msg = f"Error loading Excel file: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    def get_sheet(self, sheet_name: str):

        if sheet_name not in self.workbook.sheetnames:
            available = ', '.join(self.workbook.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook. "
                f"Available sheets: {available}"
            )
        
        return self.workbook[sheet_name]
    
    def sheet_exists(self, sheet_name: str) -> bool:

        return sheet_name in self.workbook.sheetnames
    
    def parse_period_header(self, sheet, row: int = None) -> Dict[str, Dict[str, Any]]:

        if row is None:
            row = self.PERIOD_HEADER_ROW
        
        all_period_headers = []  # ← Kumpulkan SEMUA, bukan cuma 2!
        
        logger.info(f"Parsing ALL period headers at row {row}")
        
        # Collect ALL merged cells di row 7 dengan bulan/tahun
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row == row:
                cell = sheet.cell(row, merged_range.min_col)
                header_text = str(cell.value or "").strip()
                
                if not header_text:
                    continue
                
                # Extract month and year
                month = self._extract_month(header_text)
                year = self._extract_year(header_text)
                
                # Skip jika tidak ada bulan/tahun
                if month is None or year is None:
                    logger.debug(f"Skipping '{header_text}' - no month/year")
                    continue
                
                # Tentukan tipe: 's.d.' atau 'bulan'
                is_sd = self._is_sd_header(header_text)
                
                all_period_headers.append({
                    'type': 's.d.' if is_sd else 'bulan',
                    'text': header_text,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col,
                    'month': month,
                    'year': year,
                    'is_sd': is_sd
                })
                
                logger.debug(
                    f"Found: '{header_text}' at cols {merged_range.min_col}-{merged_range.max_col}, "
                    f"Type: {'s.d.' if is_sd else 'bulan'}, Month: {month}, Year: {year}"
                )
        
        # Sort by column position
        all_period_headers.sort(key=lambda x: x['start_col'])
        
        logger.info(f"Found {len(all_period_headers)} total period headers")
        
        if not all_period_headers:
            raise ValueError(f"No period headers found at row {row}")
        
        # Return ALL headers, biar column_detector yang filter
        return {'all_periods': all_period_headers}
    
    
    def _is_sd_header(self, text: str) -> bool:

        text_lower = text.lower()
        
        # Check berbagai variasi s.d.
        sd_patterns = ['s.d.', 's.d ', 's/d', 'sd.', 'sampai dengan', 'sampai']
        
        for pattern in sd_patterns:
            if pattern in text_lower:
                return True
        
        return False
    
    def parse_type_header(self, sheet, row: int = None) -> Dict[int, str]:

        if row is None:
            row = self.TYPE_HEADER_ROW
        
        type_mapping = {}
        
        logger.info(f"Parsing type headers at row {row}")
        
        # Scan semua kolom untuk cari REAL/RKAP
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            header_text = str(cell.value or "").strip().upper()
            
            if not header_text:
                continue
            
            # Check apakah mengandung REAL atau RKAP
            if 'REAL' in header_text:
                type_mapping[col] = 'REAL'
                logger.debug(f"  Col {self._col_index_to_letter(col)}: REAL")
            elif 'RKAP' in header_text:
                type_mapping[col] = 'RKAP'
                logger.debug(f"  Col {self._col_index_to_letter(col)}: RKAP")
        
        if not type_mapping:
            raise ValueError(f"No REAL/RKAP headers found at row {row}")
        
        logger.info(f"Found {len(type_mapping)} type headers")
        
        return type_mapping
    
    def _extract_month(self, text: str) -> Optional[int]:
        
        text_lower = text.lower()
        
        # Cari semua bulan yang muncul di text
        found_months = []
        for month_key, month_num in self.MONTH_MAP.items():
            # Gunakan word boundary untuk match lebih akurat
            # Cek apakah month_key muncul sebagai kata terpisah
            if month_key in text_lower:
                # Hitung posisi kemunculan
                pos = text_lower.find(month_key)
                found_months.append((month_num, pos))
        
        if not found_months:
            logger.warning(f"Could not extract month from: {text}")
            return None
        
        # Jika ada multiple matches, ambil yang pertama muncul
        found_months.sort(key=lambda x: x[1])
        month_num = found_months[0][0]
        
        logger.debug(f"Extracted month {month_num} from: {text}")
        return month_num
    
    def _extract_year(self, text: str) -> Optional[int]:
        
        # Cari pattern 4 digit yang dimulai dengan 20
        match = re.search(r'20\d{2}', text)
        
        if match:
            return int(match.group())
        
        logger.warning(f"Could not extract year from: {text}")
        return None
    
    def _parse_period_header_fallback(self, sheet, row: int, existing_headers: Dict) -> Dict:

        logger.info("Using fallback method to parse period headers")
        
        # Scan semua kolom
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            header_text = str(cell.value or "").strip()
            
            if not header_text or len(header_text) < 5:
                continue
            
            # Extract month and year
            month = self._extract_month(header_text)
            year = self._extract_year(header_text)
            
            if month and year:
                header_lower = header_text.lower()
                
                # Determine if this is 'bulan' or 's.d.'
                if 's.d.' in header_lower or 's.d ' in header_lower:
                    if 's.d.' not in existing_headers:
                        existing_headers['s.d.'] = {
                            'text': header_text,
                            'start_col': col,
                            'end_col': col + 1,  # Assume 2 columns
                            'month': month,
                            'year': year
                        }
                        logger.info(f"  Fallback found s.d.: {header_text}")
                else:
                    if 'bulan' not in existing_headers:
                        existing_headers['bulan'] = {
                            'text': header_text,
                            'start_col': col,
                            'end_col': col + 1,  # Assume 2 columns
                            'month': month,
                            'year': year
                        }
                        logger.info(f"  Fallback found bulan: {header_text}")
        
        return existing_headers
    
    def get_cell_value(self, sheet, row: int, col: int) -> float:

        try:
            value = sheet.cell(row, col).value
            
            # Handle None
            if value is None:
                return 0.0
            
            # Handle string yang bisa diconvert ke number
            if isinstance(value, str):
                # Remove common formatting (comma, space)
                value = value.replace(',', '').replace(' ', '').strip()
                if value == '' or value == '-':
                    return 0.0
            
            # Convert to float
            return float(value)
            
        except (ValueError, TypeError, AttributeError) as e:
            logger.warning(
                f"Could not convert cell value at row {row}, col {col}: {e}. "
                f"Returning 0.0"
            )
            return 0.0
    
    def get_cell_value_raw(self, sheet, row: int, col: int) -> Any:

        return sheet.cell(row, col).value
    
    def _col_index_to_letter(self, col: int) -> str:

        result = ""
        while col > 0:
            col -= 1
            result = chr(col % 26 + 65) + result
            col //= 26
        return result
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:

        sheet = self.get_sheet(sheet_name)
        
        return {
            'name': sheet_name,
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'dimensions': sheet.dimensions,
            'has_merged_cells': len(list(sheet.merged_cells.ranges)) > 0
        }
    
    def close(self) -> None:
        """Close workbook"""
        if self.workbook:
            self.workbook.close()
            logger.info("Excel workbook closed")


if __name__ == "__main__":
    # Test excel reader
    import logging
    logging.basicConfig(level=logging.INFO)
    
    print("\n" + "="*70)
    print("EXCEL READER TEST")
    print("="*70)
    
    try:
        # Ganti dengan path file Excel yang sebenarnya
        reader = ExcelReader("data/input/test.xlsx")
        
        print(f"\nAvailable sheets: {len(reader.sheet_names)}")
        for sheet_name in reader.sheet_names:
            info = reader.get_sheet_info(sheet_name)
            print(f"  - {sheet_name}: {info['max_row']} rows x {info['max_column']} cols")
        
        reader.close()
        
    except Exception as e:
        print(f"Error: {e}")