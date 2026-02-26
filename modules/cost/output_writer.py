"""
output_writer.py
================
Module untuk menulis output ke Excel (in-memory atau disk)
"""

import pandas as pd
import io
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class OutputWriter:

    def __init__(self, output_dir: str = "data/output"):

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Output directory: {self.output_dir}")
    
    def write_excel_to_bytes(
        self, 
        data: List[Dict[str, Any]], 
        period: str
    ) -> bytes:

        logger.info(f"Writing {len(data)} rows to Excel (in-memory)")
        
        # Convert to DataFrame
        df = self._prepare_dataframe(data)
        
        # Generate sheet name dari period
        sheet_name = self._get_sheet_name_from_period(period)
        
        # Generate summary
        summary_df = self.get_summary_statistics(data)
        
        # Create BytesIO buffer (in-memory)
        buffer = io.BytesIO()
        
        # Write to Excel in memory
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Data lengkap
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet_data = writer.sheets[sheet_name]
            self._format_worksheet(worksheet_data, df)
            
            # Sheet 2: Summary
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            worksheet_summary = writer.sheets['Summary']
            self._format_worksheet(worksheet_summary, summary_df)
        
        # Get bytes
        buffer.seek(0)
        file_bytes = buffer.read()
        
        file_size_kb = len(file_bytes) / 1024
        
        logger.info(f"Excel created in memory ({file_size_kb:.2f} KB, 2 sheets)")
        print(f"✓ Excel created: {file_size_kb:.2f} KB")
        print(f"  - Sheet 1: '{sheet_name}' ({len(df):,} rows)")
        print(f"  - Sheet 2: 'Summary' ({len(summary_df):,} rows)")
        
        return file_bytes
    
    def _prepare_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:

        # Define column order
        column_order = [
            'holding',
            'kode_perusahaan',
            'periode',
            'payment_type',
            'cost_description',
            'REAL',
            'RKAP',
            'REAL_SD',
            'RKAP_SD'
        ]
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns
        df = df[column_order]
        
        # JANGAN round di sini! Data asli tetap utuh
        # Formatting akan dilakukan di Excel cell format
        
        return df
    
    def _get_sheet_name_from_period(self, period: str) -> str:

        try:
            # Parse period
            dt = datetime.strptime(period, '%d/%m/%Y')
            
            # Mapping bulan ke Bahasa Indonesia
            month_names = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
                5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
                9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            
            month_name = month_names[dt.month]
            year = dt.year
            
            return f"{month_name} {year}"
        
        except ValueError:
            # Fallback jika format tidak sesuai
            logger.warning(f"Could not parse period '{period}', using default sheet name")
            return "Data"
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame) -> None:

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Format header row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Auto-adjust column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set width dengan minimum 10 dan maximum 50
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format data cells
        for row_num in range(2, worksheet.max_row + 1):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                
                # Align text based on column type
                col_name = df.columns[col_num - 1]
                if col_name in ['REAL', 'RKAP', 'REAL_SD', 'RKAP_SD']:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Format angka: 2 desimal, separator ribuan
                    # Data tetap full precision, hanya tampilan yang diformat
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Freeze top row
        worksheet.freeze_panes = 'A2'
        
        logger.debug("Worksheet formatting applied")
    
    def get_summary_statistics(self, data: List[Dict[str, Any]]) -> pd.DataFrame:

        df = pd.DataFrame(data)
        
        # Group by holding and payment_type
        summary = df.groupby(['holding', 'payment_type']).agg({
            'REAL': 'sum',
            'RKAP': 'sum',
            'REAL_SD': 'sum',
            'RKAP_SD': 'sum'
        }).round(2)
        
        # Reset index untuk jadi kolom biasa
        summary = summary.reset_index()
        
        # Rename columns untuk lebih jelas
        summary.columns = [
            'Holding', 
            'Payment Type', 
            'REAL (Bulan Ini)', 
            'RKAP (Bulan Ini)', 
            'REAL (s.d.)', 
            'RKAP (s.d.)'
        ]
        
        # Add total row
        total_row = pd.DataFrame([{
            'Holding': 'TOTAL',
            'Payment Type': 'ALL',
            'REAL (Bulan Ini)': summary['REAL (Bulan Ini)'].sum(),
            'RKAP (Bulan Ini)': summary['RKAP (Bulan Ini)'].sum(),
            'REAL (s.d.)': summary['REAL (s.d.)'].sum(),
            'RKAP (s.d.)': summary['RKAP (s.d.)'].sum()
        }])
        
        summary = pd.concat([summary, total_row], ignore_index=True)
        
        return summary


if __name__ == "__main__":
    # Test output writer
    import logging
    logging.basicConfig(level=logging.INFO)
    
    print("\n" + "="*70)
    print("OUTPUT WRITER TEST")
    print("="*70)
    
    # Sample data
    sample_data = [
        {
            'holding': 'PTPN III',
            'kode_perusahaan': 'CHOL',
            'periode': '2024-12',
            'payment_type': 'CASH',
            'cost_description': 'Biaya Dewan Komisaris',
            'REAL': 1000000.50,
            'RKAP': 1200000.00,
            'REAL_SD': 12000000.50,
            'RKAP_SD': 14400000.00
        },
        {
            'holding': 'PTPN III',
            'kode_perusahaan': 'CHOL',
            'periode': '2024-12',
            'payment_type': 'CASH',
            'cost_description': 'Biaya Direksi',
            'REAL': 2000000.75,
            'RKAP': 2500000.00,
            'REAL_SD': 24000000.75,
            'RKAP_SD': 30000000.00
        }
    ]
    
    try:
        writer = OutputWriter()
        
        # Test in-memory Excel generation
        file_bytes = writer.write_excel_to_bytes(sample_data, "31/12/2024")
        print(f"\n✓ Generated {len(file_bytes)} bytes")
        
        # Optional: Save to disk for testing
        with open("test_output.xlsx", "wb") as f:
            f.write(file_bytes)
        print("✓ Test file saved: test_output.xlsx")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()