import openpyxl
from typing import Optional, Dict

from modules.demografi.utils import get_logger, safe_int, FileError


# Get logger instance
logger = get_logger()


class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def __enter__(self):
        try:
            # Load workbook
            # data_only=True: Baca HASIL formula, bukan formula-nya
            # read_only=True: Faster, tapi tidak bisa edit
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=True,
                read_only=True
            )
            logger.debug(f"Workbook opened: {self.file_path}")
            return self
        except Exception as e:
            raise FileError(f"Tidak bisa membuka file: {str(e)}")
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.close()
            logger.debug("Workbook closed")
    
    def select_sheet(self, sheet_name: str):
        if sheet_name not in self.workbook.sheetnames:
            available = ', '.join(self.workbook.sheetnames)
            raise FileError(
                f"Sheet '{sheet_name}' tidak ditemukan. "
                f"Available: {available}"
            )
        
        self.worksheet = self.workbook[sheet_name]
        logger.debug(f"Sheet selected: {sheet_name}")
    
    def read_cell(self, cell_address: str) -> int:
        """
        Baca nilai dari cell.
        
        Handle:
        - Cell kosong (None) → 0
        - String kosong ('') → 0
        - Error (#REF!, #VALUE!) → 0
        - Valid number → convert ke int
        
        Args:
            cell_address: Cell address (ex: 'C10', 'AB5')
        
        Returns:
            int: Nilai cell (0 jika error/kosong)
            
        Example:
            >>> value = reader.read_cell('C10')
            >>> print(value)  # 1
            
            >>> value = reader.read_cell('C11')  # Cell kosong
            >>> print(value)  # 0
        """
        if not self.worksheet:
            raise FileError("Sheet belum dipilih. Gunakan select_sheet() dulu")
        
        try:
            # Baca cell value
            value = self.worksheet[cell_address].value
            
            # Convert ke int (safe)
            result = safe_int(value, default=0)
            
            # Logging untuk debugging
            if value is None or value == '':
                logger.debug(f"Cell {cell_address}: kosong → 0")
            else:
                logger.debug(f"Cell {cell_address}: {value} → {result}")
            
            return result
            
        except Exception as e:
            logger.warning(f"Error reading cell {cell_address}: {e} → 0")
            return 0
    
    def read_cells(self, cell_mapping: Dict[str, str]) -> Dict[str, int]:
        """
        Baca multiple cells sekaligus.
        Lebih efficient daripada read_cell() satu-satu.
        
        Args:
            cell_mapping: Dict {kategori: cell_address}
        
        Returns:
            dict: {kategori: nilai}
            
        Example:
            >>> mapping = {
            ...     '<26': 'C10',
            ...     '26-30': 'C11',
            ...     '31-35': 'C12'
            ... }
            >>> results = reader.read_cells(mapping)
            >>> print(results)
            {'<26': 1, '26-30': 31, '31-35': 45}
        """
        results = {}
        for kategori, cell in cell_mapping.items():
            results[kategori] = self.read_cell(cell)
        return results
    
    def get_sheet_names(self) -> list:
        """
        Get list of all sheet names di workbook.
        
        Returns:
            list: Sheet names
            
        Example:
            >>> sheets = reader.get_sheet_names()
            >>> print(sheets)
            ['SGN', 'PTPN3', 'Summary']
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames


def read_dimension_data(file_path: str, sheet_name: str, 
                       cell_mapping: Dict[str, str]) -> Dict[str, int]:
    """
    Helper function untuk baca data 1 dimensi.
    Standalone function (tidak perlu manage ExcelReader manually).
    
    Args:
        file_path: Path ke Excel file
        sheet_name: Nama sheet
        cell_mapping: Dict {kategori: cell_address}
    
    Returns:
        dict: {kategori: jumlah}
        
    Example:
        >>> mapping = {'<26': 'C10', '26-30': 'C11'}
        >>> data = read_dimension_data('data.xlsx', 'SGN', mapping)
        >>> print(data)
        {'<26': 1, '26-30': 31}
    """
    with ExcelReader(file_path) as reader:
        reader.select_sheet(sheet_name)
        return reader.read_cells(cell_mapping)


def get_available_sheets(file_path: str) -> list:
    """
    Get list of available sheets in Excel.
    Standalone function.
    
    Args:
        file_path: Path ke Excel file
    
    Returns:
        list: Sheet names
        
    Example:
        >>> sheets = get_available_sheets('data.xlsx')
        >>> print(sheets)
        ['SGN', 'PTPN3', 'Summary']
    """
    try:
        with ExcelReader(file_path) as reader:
            return reader.get_sheet_names()
    except Exception as e:
        logger.error(f"Error getting sheets: {e}")
        return []