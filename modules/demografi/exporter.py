import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict

from modules.demografi.utils import get_logger, ensure_dir


logger = get_logger()


def export_to_excel(dataframes: Dict[str, pd.DataFrame], 
                   output_dir: str,
                   original_filename: str,                    
                   perusahaan: str = None) -> str:
    try:
        
        ensure_dir(output_dir)

        orig_name = Path(original_filename).stem   # data_sdm_ptpn3
        ext = Path(original_filename).suffix        # .xlsx

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        company_part = f"_{perusahaan}" if perusahaan else ""

        output_filename = f"{orig_name}_converted{company_part}_{timestamp}{ext}"
        output_path = str(Path(output_dir) / output_filename)

        logger.info(f"📄 Exporting ke: {output_path}")
        
       
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Mapping dimensi ke sheet name
            sheet_mapping = {
                'gender': 'Data Gender',
                'pendidikan': 'Data Pendidikan',
                'usia': 'Data Usia',
                'unit_kerja': 'Data Unit Kerja',
                'tren': 'Data Tren'
            }
            
            # Export setiap dimensi ke sheet-nya
            for dimensi, sheet_name in sheet_mapping.items():
                if dimensi in dataframes:
                    df = dataframes[dimensi]
                    
                    # Untuk tren, column order berbeda (4 kolom saja)
                    if dimensi == 'tren':
                        # Pastikan column order: periode, nama_perusahaan, kategori, jumlah
                        df = df[['periode', 'nama_perusahaan', 'kategori', 'jumlah']]
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"   ✅ Sheet '{sheet_name}': {len(df)} rows")
            
            # === Sheet: Summary ===
            summary_data = []
            
            for dimensi, df in dataframes.items():
                # Overall total per dimensi
                total = df['jumlah'].sum()
                count = len(df)
                
                summary_data.append({
                    'Dimensi': dimensi.upper(),
                    'Total Karyawan': total,
                    'Jumlah Baris': count
                })
                
                # Group by kelompok (jika ada kolom kelompok)
                if 'kelompok' in df.columns:
                    by_kelompok = df.groupby('kelompok')['jumlah'].sum().to_dict()
                    for kelompok, jumlah in by_kelompok.items():
                        summary_data.append({
                            'Dimensi': f"  → {kelompok}",
                            'Total Karyawan': jumlah,
                            'Jumlah Baris': ''
                        })
                elif dimensi == 'tren':
                    # Untuk tren, group by kategori (TETAP/NON TETAP)
                    by_kategori = df.groupby('kategori')['jumlah'].sum().to_dict()
                    for kategori, jumlah in by_kategori.items():
                        summary_data.append({
                            'Dimensi': f"  → {kategori}",
                            'Total Karyawan': jumlah,
                            'Jumlah Baris': ''
                        })
            
            if summary_data:
                summary = pd.DataFrame(summary_data)
                summary.to_excel(writer, sheet_name='Summary', index=False)
                logger.info(f"   ✅ Sheet 'Summary': {len(summary)} rows")
        
        
        _format_excel(output_path)
        
        logger.info(f"✅ Export berhasil: {output_path}")
        return str(output_path)
        
    except Exception as e:
        logger.error(f"❌ Error export: {e}")
        raise


def _format_excel(file_path: str):
    """Apply formatting ke Excel file."""
    try:
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # --- Header Styling ---
            header_fill = PatternFill(
                start_color="366092", 
                end_color="366092", 
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )
            
            # --- Auto-adjust Column Width ---
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # --- Freeze Header Row ---
            ws.freeze_panes = 'A2'
            
            # --- Number Format untuk kolom 'jumlah' / 'Total Karyawan' ---
            # Find jumlah column
            jumlah_cols = []
            for idx, cell in enumerate(ws[1], 1):
                if cell.value in ['jumlah', 'Total Karyawan']:
                    jumlah_cols.append(idx)
            
            # Apply number format
            for col_idx in jumlah_cols:
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
        
        wb.save(file_path)
        logger.debug("Excel formatting applied")
        
    except Exception as e:
        logger.warning(f"Error formatting Excel: {e}")